from abc import ABC,abstractmethod
from config import Config
from openpyxl import Workbook,styles
from openpyxl.chart import BarChart,Reference
from openpyxl.styles import Border,Side
from api import db
from api.models import UserAnswer,User
from sqlalchemy import func

border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))


class CreateFile(ABC):
    def __init__(self):
        self.media = Config.MEDIA
    @abstractmethod
    def create(self):
        pass

class CreateFileExcel(CreateFile):
    def __init__(self):
        super().__init__()
    def create(self):
        wb = Workbook()
        wb.create_sheet(title='Результат', index=0)
        sheet = wb['Результат']
        user_count_db = db.session.query(User).with_entities(func.count()).filter(User.is_test == True,User.is_admin == False).first()
        sheet.merge_cells(f"A{sheet.cell(row=1, column=1).row}:C{sheet.cell(row=1, column=3).row}")
        sheet.cell(row=1, column=1).value = f'Количество тестируемых: {user_count_db[0]}'
        result_test_db = db.session.query(UserAnswer).with_entities(UserAnswer.test_id,UserAnswer.is_right,func.count()).group_by(UserAnswer.test_id,UserAnswer.is_right).all()
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.cell(row=2, column=1).value = 'Вопрос'
        sheet.cell(row=2, column=2).value = 'Правильно'
        sheet.cell(row=2, column=3).value = 'Не правильно'
        for i in range(3):
            sheet.cell(row=2, column=1 + i).border = border
            sheet.cell(row=2, column=1 + i).alignment = styles.Alignment(horizontal="center", vertical="center")

        row = 2
        for result in result_test_db:
            row +=1
            if result[0] == sheet.cell(row=row-1, column=1).value:
                row -= 1
            sheet.cell(row=row, column=1).value = result[0]
            if result[1]:
                col = 2
            else:
                col = 3
            sheet.cell(row=row, column=col).value = result[2]
            for i in range(3):
                sheet.cell(row=row, column=1+i).border = border
                sheet.cell(row=row, column=1+i).alignment = styles.Alignment(horizontal="center", vertical="center")

        values = Reference(worksheet=sheet,
                           min_row=2,
                           max_row=row,
                           min_col=2,
                           max_col=3)

        chart = BarChart()
        chart.type = "col"
        chart.style = 12
        chart.grouping = "stacked"
        chart.overlap = 100
        chart.title = 'Результаты теста'

        chart.add_data(values, titles_from_data=True)

        sheet.add_chart(chart, "E3")
        chart.width = 20
        chart.height = 5

        wb.save(self.media+'report.xlsx')
        return self.media+'report.xlsx'
